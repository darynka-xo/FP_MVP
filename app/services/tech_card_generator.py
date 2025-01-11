import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import tempfile


class TechCardGenerator:
    def __init__(self, template_path):
        self.template_path = template_path

        self.green_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
        self.yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.thin_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')

    def write_to_cell(self, ws, cell, value, style_config=None):
        """Write value to a cell"""
        try:
            ws[cell].value = value
            if style_config:
                if 'fill' in style_config:
                    ws[cell].fill = style_config['fill']
                if 'border' in style_config:
                    ws[cell].border = style_config['border']
                if 'alignment' in style_config:
                    ws[cell].alignment = style_config['alignment']
        except Exception as e:
            print(f"Error writing to cell {cell}: {str(e)}")

    def generate_tech_card(self, part1_data, part2_data):
        try:
            wb = load_workbook(self.template_path)
            ws = wb.active

            # Write to the main cells of merged ranges (using first cell in range)
            merged_cells = {
                'D4': datetime.now().strftime('%d/%m/%Y'),  # Date
                'C6': part1_data.customer,  # Customer
                'H6': part1_data.design,  # Design
                'C11': part1_data.product_type,  # Product type
                'C17': part1_data.manufacturer,  # Manufacturer
                'I17': part1_data.meters_per_circulation,  # Meters per circulation
                'K17': part1_data.kg_per_circulation,  # KG per circulation
                'I18': part1_data.bottom_material_meters,  # Bottom material meters
                'K18': part1_data.bottom_material_kg,  # Bottom material kg
                'F34': part1_data.bottom_width  # Bottom width
            }

            # Write to merged cells
            for cell, value in merged_cells.items():
                self.write_to_cell(ws, cell, value, {
                    'alignment': self.center_alignment,
                    'border': self.thin_border
                })

            # Write to single cells
            single_cells = {
                'E6': part1_data.order_number,
                'F6': part1_data.circulation,
                'G6': part1_data.cup_article,
                'E12': part1_data.throat_diameter,
                'F12': part1_data.bottom_diameter,
                'G12': part1_data.height,
                'H12': part1_data.capacity,
                'E17': part1_data.name,
                'F17': part1_data.density,
                'G17': part1_data.width,
                'H17': part1_data.pe_layer,
                'D18': part1_data.quantity_per_rapport,
                'E21': part1_data.sleeve,
                'C34': part1_data.tooling_number,
                'D34': part1_data.quantity_per_rapport,
                'I34': part1_data.glasses_per_sleeve,
                'J34': part1_data.sleeves_per_box,
                'K34': part1_data.corrugated_box_size
            }

            for cell, value in single_cells.items():
                self.write_to_cell(ws, cell, value, {
                    'alignment': self.center_alignment,
                    'border': self.thin_border
                })

            # Write Part 2 data
            if part2_data:
                tech_card_part2_cells = {
                    'E24': part2_data.lineature_anilox,
                    'E25': part2_data.shaft_number,
                    'E26': part2_data.name,
                    'E27': part2_data.color,
                    'E28': part2_data.viscosity,
                    'E29': part2_data.consumption,
                    'E30': part2_data.comments
                }

                for cell, value in tech_card_part2_cells.items():
                    self.write_to_cell(ws, cell, value, {
                        'alignment': self.center_alignment,
                        'border': self.thin_border
                    })

            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)

            with open(temp_file.name, 'rb') as f:
                excel_content = f.read()

            os.unlink(temp_file.name)
            return excel_content

        except Exception as e:
            print(f"Error in generate_tech_card: {str(e)}")
            raise