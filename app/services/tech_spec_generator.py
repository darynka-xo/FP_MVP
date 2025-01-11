import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import tempfile

class TechSpecGenerator:
    def __init__(self, template_path):
        self.template_path = template_path

    def write_to_cell(self, ws, cell, value):
        """Write value to a cell with proper formatting"""
        try:
            ws[cell].value = value
        except Exception as e:
            print(f"Error writing to cell {cell}: {str(e)}")

    def generate_tech_spec(self, spec_data):
        try:
            wb = load_workbook(self.template_path)
            ws = wb.active

            # Map data to cells (column B)
            cell_mappings = {
                'B1': spec_data.order_number,
                'B2': spec_data.article,
                'B3': spec_data.design,
                'B4': spec_data.production_start_date,
                'B5': spec_data.cup_type,
                'B6': spec_data.order_quantity,
                'B7': spec_data.color,
                'B8': spec_data.total_cup_weight,
                'B9': spec_data.machine,
                'B10': spec_data.material,
                'B12': spec_data.paper_density,
                'B13': spec_data.side_wall,
                'B14': spec_data.size_mm,
                'B15': spec_data.total_density,
                'B16': spec_data.bottom,
                'B17': spec_data.size_mm_82_72,
                'B18': spec_data.total_density_82_72,
                'B20': spec_data.strokes_per_min,
                'B21': spec_data.roll_width_after_print,
                'B23': spec_data.side_wall_weight_per_cup,
                'B24': spec_data.side_wall_cutting_weight,
                'B25': spec_data.total_side_wall_weight,
                'B26': spec_data.blanker_waste_norm,
                'B27': spec_data.number_of_streams,
                'B28': spec_data.utilization_ratio,
                'B29': spec_data.productivity_per_hour,
                'B31': spec_data.blank_consumption_per_1000,
                'B32': spec_data.side_wall_blanks,
                'B34': spec_data.bottom_weight,
                'B35': spec_data.bottom_cutting_weight,
                'B37': spec_data.pe_packaging,
                'B38': spec_data.pe_weight,
                'B39': spec_data.pe_packaging_consumption,
                'B40': spec_data.stacks_per_box,
                'B41': spec_data.items_per_stack,
                'B42': spec_data.items_per_box
            }

            # Write values to cells
            for cell, value in cell_mappings.items():
                self.write_to_cell(ws, cell, value)

            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)

            with open(temp_file.name, 'rb') as f:
                excel_content = f.read()

            os.unlink(temp_file.name)
            return excel_content

        except Exception as e:
            print(f"Error in generate_tech_spec: {str(e)}")
            raise